# C:\office_crm\views\coverage_report.py

from pathlib import Path
from datetime import datetime
import html

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from utils.customer import get_customers
from tools.coverage_pdf_parser import parse_coverage_pdf
from tools.coverage_template_writer import write_coverage_report
from tools.ollama_client import (
    is_ollama_available,
    analyze_coverage_with_ollama,
)


BASE_DIR = Path(__file__).resolve().parent.parent

UPLOAD_ROOT = BASE_DIR / "uploads" / "coverage_report"
OUTPUT_ROOT = BASE_DIR / "outputs" / "coverage_report"

TEMPLATE_PATH = BASE_DIR / "templates" / "간단하게 보장분석표 - 복사본.xlsx"


def safe_name(value):
    text = str(value or "").strip()
    for ch in r'\/:*?"<>|':
        text = text.replace(ch, "_")
    return text or "unknown"


def save_uploaded_file(uploaded_file, save_dir):
    save_dir.mkdir(parents=True, exist_ok=True)
    save_path = save_dir / safe_name(uploaded_file.name)

    with open(save_path, "wb") as f:
        f.write(uploaded_file.getbuffer())

    return save_path


def color_to_hex(color):
    if not color:
        return None

    try:
        if color.type == "rgb" and color.rgb:
            value = color.rgb
            if len(value) == 8:
                return f"#{value[2:]}"
            if len(value) == 6:
                return f"#{value}"
    except Exception:
        return None

    return None


def cell_css(cell):
    css = []

    fill_color = color_to_hex(cell.fill.fgColor)
    if fill_color and fill_color.upper() != "#000000":
        css.append(f"background:{fill_color}")

    font_color = color_to_hex(cell.font.color)
    if font_color:
        css.append(f"color:{font_color}")

    if cell.font.bold:
        css.append("font-weight:700")

    if cell.font.sz:
        css.append(f"font-size:{int(cell.font.sz)}px")

    if cell.alignment.horizontal:
        css.append(f"text-align:{cell.alignment.horizontal}")

    if cell.alignment.vertical:
        css.append(f"vertical-align:{cell.alignment.vertical}")

    if cell.alignment.wrap_text:
        css.append("white-space:normal")
    else:
        css.append("white-space:nowrap")

    css.append("border:1px solid #222")
    css.append("padding:2px 4px")
    css.append("height:22px")
    css.append("box-sizing:border-box")

    return ";".join(css)


def get_merged_map(ws):
    merged_map = {}
    skip_cells = set()

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds

        rowspan = max_row - min_row + 1
        colspan = max_col - min_col + 1

        merged_map[(min_row, min_col)] = {
            "rowspan": rowspan,
            "colspan": colspan,
        }

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if not (row == min_row and col == min_col):
                    skip_cells.add((row, col))

    return merged_map, skip_cells


def worksheet_to_html(xlsx_path, sheet_name=None, max_rows=120, max_cols=40):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    merged_map, skip_cells = get_merged_map(ws)

    max_row = min(ws.max_row, max_rows)
    max_col = min(ws.max_column, max_cols)

    colgroup = []

    for col_idx in range(1, max_col + 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        width = ws.column_dimensions[letter].width or 10
        px = max(45, min(int(width * 7), 190))
        colgroup.append(f"<col style='width:{px}px'>")

    rows_html = []

    for row_idx in range(1, max_row + 1):
        row_cells = []

        for col_idx in range(1, max_col + 1):
            if (row_idx, col_idx) in skip_cells:
                continue

            cell = ws.cell(row=row_idx, column=col_idx)

            if isinstance(cell, MergedCell):
                continue

            value = cell.value
            display_value = "" if value is None else str(value)

            attrs = ""

            if (row_idx, col_idx) in merged_map:
                info = merged_map[(row_idx, col_idx)]

                if info["rowspan"] > 1:
                    attrs += f" rowspan='{info['rowspan']}'"

                if info["colspan"] > 1:
                    attrs += f" colspan='{info['colspan']}'"

            row_cells.append(
                f"<td{attrs} style='{cell_css(cell)}'>"
                f"{html.escape(display_value)}"
                f"</td>"
            )

        rows_html.append(f"<tr>{''.join(row_cells)}</tr>")

    return f"""
    <html>
    <head>
    <style>
        body {{
            margin:0;
            background:#f4f6f8;
            font-family:Malgun Gothic, Apple SD Gothic Neo, Arial, sans-serif;
        }}

        .excel-wrap {{
            width:100%;
            height:900px;
            overflow:auto;
            background:#f4f6f8;
            border:1px solid #d0d5dd;
            border-radius:10px;
        }}

        .excel-table {{
            border-collapse:collapse;
            table-layout:fixed;
            background:white;
            margin:12px;
            min-width:2100px;
        }}

        .excel-table td {{
            line-height:1.25;
        }}

        @media print {{
            body {{
                background:white;
            }}

            .excel-wrap {{
                height:auto;
                overflow:visible;
                border:none;
            }}

            .excel-table {{
                margin:0;
            }}
        }}
    </style>
    </head>
    <body>
        <div class="excel-wrap">
            <table class="excel-table">
                <colgroup>
                    {''.join(colgroup)}
                </colgroup>
                <tbody>
                    {''.join(rows_html)}
                </tbody>
            </table>
        </div>
    </body>
    </html>
    """


def render_result_preview(output_path):
    st.subheader("보장분석표 미리보기")

    st.caption(
        "업로드한 엑셀 첫 번째 시트 형식을 웹에서 확인하는 화면입니다. "
        "좌우 스크롤로 전체 표를 볼 수 있습니다."
    )

    try:
        html_view = worksheet_to_html(
            output_path,
            sheet_name=None,
            max_rows=120,
            max_cols=40,
        )
        components.html(html_view, height=920, scrolling=True)

    except Exception as e:
        st.error(f"보장분석표 미리보기 생성 실패: {e}")


def render_extracted_data(parsed_data):
    tab1, tab2, tab3 = st.tabs(["자동분류 요약", "계약 리스트", "담보 상세"])

    with tab1:
        summary = parsed_data.get("summary") or []
        if summary:
            st.dataframe(
                pd.DataFrame(summary),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.info("자동분류 요약 데이터가 없습니다.")

    with tab2:
        contracts = parsed_data.get("contracts") or []
        if contracts:
            st.dataframe(
                pd.DataFrame(contracts),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.info("계약 리스트 추출 데이터가 없습니다.")

    with tab3:
        coverages = parsed_data.get("coverages") or []
        if coverages:
            st.dataframe(
                pd.DataFrame(coverages),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.info("담보 상세 추출 데이터가 없습니다.")


def render_ai_analysis(parsed_data):
    st.subheader("AI 보장분석 코멘트")

    st.caption(
        "Ollama 로컬 AI를 사용합니다. "
        "AI 결과는 참고용이며, 최종 판단은 약관과 원본 증권 확인이 필요합니다."
    )

    if not is_ollama_available():
        st.warning(
            "Ollama 서버에 연결할 수 없습니다. "
            "CMD에서 `ollama run qwen2.5:7b` 실행 후 다시 시도하세요."
        )
        return

    model = st.selectbox(
        "사용할 Ollama 모델",
        ["qwen2.5:7b", "qwen2.5:14b", "llama3.1:8b"],
        index=0,
    )

    if st.button("AI 보장분석 실행", use_container_width=True):
        with st.spinner("Ollama가 보장 내용을 분석 중입니다..."):
            ai_result = analyze_coverage_with_ollama(
                parsed_data=parsed_data,
                model=model,
            )

        st.session_state.last_coverage_ai_result = ai_result

    ai_result = st.session_state.get("last_coverage_ai_result")

    if ai_result:
        st.markdown(ai_result)
    else:
        st.info("AI 분석 실행 버튼을 누르면 이곳에 분석 코멘트가 표시됩니다.")


def coverage_report_page(user):
    st.subheader("보장분석표 생성")

    customers = get_customers(user)

    if not customers:
        st.info("먼저 고객을 등록해야 보장분석표를 생성할 수 있습니다.")
        return

    if "selected_coverage_customer_id" not in st.session_state:
        st.session_state.selected_coverage_customer_id = None

    customer_options = {"(선택)": None}

    for customer in customers:
        label = (
            f"{customer.get('name') or '이름없음'} / "
            f"{customer.get('phone') or '연락처 없음'}"
        )
        customer_options[label] = customer["id"]

    option_labels = list(customer_options.keys())
    customer_ids = list(customer_options.values())

    default_index = 0
    if st.session_state.selected_coverage_customer_id in customer_ids:
        default_index = customer_ids.index(
            st.session_state.selected_coverage_customer_id
        )

    selected_label = st.selectbox(
        "고객 선택",
        option_labels,
        index=default_index,
    )

    selected_customer_id = customer_options[selected_label]
    st.session_state.selected_coverage_customer_id = selected_customer_id

    if selected_customer_id is None:
        st.info("보장분석표를 생성할 고객을 선택하세요.")
        return

    selected_customer_name = selected_label.split(" / ")[0]

    st.divider()

    left, right = st.columns([2, 1])

    with left:
        uploaded_pdf = st.file_uploader(
            "보장분석 PDF 업로드",
            type=["pdf"],
        )

    with right:
        output_name = st.text_input(
            "결과 파일명",
            value=f"{safe_name(selected_customer_name)}_보장분석표.xlsx",
        )

        use_template = st.checkbox(
            "엑셀 첫 번째 시트 템플릿 사용",
            value=True,
        )

    if use_template and not TEMPLATE_PATH.exists():
        st.warning(
            "템플릿 파일이 없습니다.\n\n"
            f"아래 위치에 넣어주세요.\n\n{TEMPLATE_PATH}"
        )

    run = st.button(
        "보장분석표 생성",
        use_container_width=True,
        type="primary",
    )

    if run:
        if not uploaded_pdf:
            st.warning("PDF 파일을 업로드하세요.")
            return

        job_time = datetime.now().strftime("%Y%m%d_%H%M%S")

        job_upload_dir = (
            UPLOAD_ROOT
            / f"customer_{selected_customer_id}"
            / job_time
        )

        job_output_dir = (
            OUTPUT_ROOT
            / f"customer_{selected_customer_id}"
            / job_time
        )

        try:
            saved_pdf_path = save_uploaded_file(
                uploaded_pdf,
                job_upload_dir,
            )

            with st.spinner("PDF에서 계약/담보 정보를 추출 중입니다..."):
                parsed_data = parse_coverage_pdf(saved_pdf_path)

            with st.spinner("엑셀 첫 번째 시트 형식으로 보장분석표를 생성 중입니다..."):
                result_path = write_coverage_report(
                    parsed_data=parsed_data,
                    output_dir=job_output_dir,
                    template_path=TEMPLATE_PATH if use_template else None,
                    output_filename=output_name.strip(),
                )

            st.session_state.last_coverage_report_result = {
                "success": True,
                "output_path": str(result_path),
                "parsed_data": parsed_data,
            }

            st.session_state.last_coverage_ai_result = None

            st.success("보장분석표 생성이 완료되었습니다.")

        except Exception as e:
            st.session_state.last_coverage_report_result = {
                "success": False,
                "message": str(e),
            }
            st.error(f"생성 중 오류가 발생했습니다: {e}")

    result = st.session_state.get("last_coverage_report_result")

    if not result:
        return

    st.divider()

    if not result.get("success"):
        st.error(result.get("message", "생성 실패"))
        return

    output_path = Path(result["output_path"])
    parsed_data = result.get("parsed_data", {})

    view_tab, ai_tab, data_tab, download_tab = st.tabs(
        [
            "보장분석표 화면",
            "AI 보장분석",
            "추출 데이터 확인",
            "출력 / 다운로드",
        ]
    )

    with view_tab:
        if output_path.exists():
            render_result_preview(output_path)
        else:
            st.error("생성된 엑셀 파일을 찾을 수 없습니다.")

    with ai_tab:
        render_ai_analysis(parsed_data)

    with data_tab:
        render_extracted_data(parsed_data)

    with download_tab:
        if output_path.exists():
            with open(output_path, "rb") as f:
                st.download_button(
                    label="보장분석표 엑셀 다운로드",
                    data=f,
                    file_name=output_path.name,
                    mime=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                    use_container_width=True,
                )

            st.code(str(output_path), language="text")
        else:
            st.error("다운로드할 엑셀 파일을 찾을 수 없습니다.")