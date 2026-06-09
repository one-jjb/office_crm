# coverage_template_writer.py
# -*- coding: utf-8 -*-

from pathlib import Path
from datetime import datetime
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = BASE_DIR / "templates"


def get_default_template_path():
    return TEMPLATE_DIR / "간단하게 보장분석표.xlsx"


def safe_sheet_name(name):
    text = str(name or "Sheet")
    for ch in "[]:*?/\\":
        text = text.replace(ch, "_")
    return text[:31]


def add_or_replace_sheet(wb, sheet_name):
    sheet_name = safe_sheet_name(sheet_name)
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    return wb.create_sheet(sheet_name)


def style_sheet(ws):
    header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, ws.max_column + 1):
        max_len = 10
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

    ws.freeze_panes = "A2"


def write_dataframe(ws, df):
    if df.empty:
        ws.append(["데이터 없음"])
        return
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    style_sheet(ws)


def create_workbook_from_template(template_path=None):
    template_path = Path(template_path) if template_path else get_default_template_path()
    if template_path.exists():
        return load_workbook(template_path)
    return Workbook()


def normalize_date_for_fill_master(value):
    text = str(value or "").strip()
    if not text:
        return ""
    return text.replace("-", ".")


def amount_to_manh_text(row):
    amount_text = str(row.get("amount") or "").strip()
    amount_number = row.get("amount_number", 0) or 0

    if amount_text:
        if amount_text.endswith("만원"):
            return amount_text
        if amount_text.endswith("원"):
            try:
                return f"{int(amount_number) // 10000}만원"
            except Exception:
                return amount_text
        return amount_text

    try:
        return f"{int(amount_number) // 10000}만원"
    except Exception:
        return ""


def group_coverages_by_contract(coverages):
    grouped = defaultdict(list)
    for row in coverages:
        insurer = str(row.get("insurer") or "").strip()
        product = str(row.get("product_name") or row.get("product") or "").strip()
        contract_date = str(row.get("contract_date") or "").strip()
        premium = str(row.get("premium") or "").strip()
        key = (insurer, product, contract_date, premium)
        grouped[key].append(row)
    return grouped


def write_extract_data_sheets(wb, parsed_data):
    """
    fill_master_with_mapping.py가 읽을 수 있는 Extract_Data_NN 시트를 생성한다.

    필요한 구조:
    - B열: 항목명 / C열: 값
    - B2:C6 = 보험사, 상품명, 계약일자, 납입기간, 보험료
    - B10 이후 = 탭 구분 다중 라인: 분류\t담보명\t가입금액\t상태
    """
    coverages = parsed_data.get("coverages") or []
    grouped = group_coverages_by_contract(coverages)

    # 기존 Extract_Data 시트 삭제
    for sname in list(wb.sheetnames):
        if str(sname).startswith("Extract_Data"):
            wb.remove(wb[sname])

    if not grouped:
        ws = add_or_replace_sheet(wb, "Extract_Data_01")
        ws["B2"] = "보험사"
        ws["C2"] = ""
        ws["B3"] = "상품명"
        ws["C3"] = ""
        ws["B4"] = "계약일자"
        ws["C4"] = ""
        ws["B5"] = "납입기간"
        ws["C5"] = ""
        ws["B6"] = "보험료"
        ws["C6"] = ""
        ws["B10"] = ""
        return

    for idx, ((insurer, product, contract_date, premium), rows) in enumerate(grouped.items(), start=1):
        ws = add_or_replace_sheet(wb, f"Extract_Data_{idx:02d}")

        ws["B2"] = "보험사"
        ws["C2"] = insurer
        ws["B3"] = "상품명"
        ws["C3"] = product
        ws["B4"] = "계약일자"
        ws["C4"] = normalize_date_for_fill_master(contract_date)
        ws["B5"] = "납입기간"
        ws["C5"] = ""
        ws["B6"] = "보험료"
        ws["C6"] = premium

        lines = []
        for row in rows:
            src_cat = str(row.get("mapped_category") or row.get("category") or "미분류").strip()
            src_name = str(row.get("coverage_name") or "").strip()
            src_amount = amount_to_manh_text(row)
            src_status = str(row.get("pay_type") or "").strip()
            if not src_name and not src_amount:
                continue
            lines.append(f"{src_cat}\t{src_name}\t{src_amount}\t{src_status}")

        ws["B10"] = "\n".join(lines)
        ws["B10"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 35
        ws.row_dimensions[10].height = max(80, min(900, len(lines) * 18))


def write_coverage_report(parsed_data, output_dir, template_path=None, output_filename=None):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    customer = parsed_data.get("customer", {})
    contracts = parsed_data.get("contracts", [])
    coverages = parsed_data.get("coverages", [])
    summary = parsed_data.get("summary", [])

    wb = create_workbook_from_template(template_path)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])

    # fill_master_with_mapping.py 입력용 시트
    write_extract_data_sheets(wb, parsed_data)

    # 웹/검수용 시트
    contracts_df = pd.DataFrame(contracts)
    coverages_df = pd.DataFrame(coverages)
    summary_df = pd.DataFrame(summary)
    customer_df = pd.DataFrame([customer]) if customer else pd.DataFrame()

    ws_customer = add_or_replace_sheet(wb, "고객정보")
    write_dataframe(ws_customer, customer_df)

    ws_contracts = add_or_replace_sheet(wb, "계약리스트_추출")
    write_dataframe(ws_contracts, contracts_df)

    ws_coverages = add_or_replace_sheet(wb, "담보상세_추출")
    write_dataframe(ws_coverages, coverages_df)

    ws_summary = add_or_replace_sheet(wb, "자동분류_요약")
    write_dataframe(ws_summary, summary_df)

    # 템플릿 주요 시트가 있을 경우 앞쪽에 요약값 입력
    for sheet_name in ("보장분석 비교", "MASTER_DATA", "Master_Data"):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws["A1"] = ws["A1"].value or "보장분석 비교"
            ws["A2"] = f"고객명: {customer.get('customer_name', '')}"
            ws["A3"] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            break

    if output_filename:
        final_name = output_filename
        if not final_name.lower().endswith(".xlsx"):
            final_name += ".xlsx"
    else:
        customer_name = customer.get("customer_name") or "고객"
        final_name = f"{customer_name}_보장분석표_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    output_path = output_dir / final_name
    wb.save(output_path)
    return str(output_path)
