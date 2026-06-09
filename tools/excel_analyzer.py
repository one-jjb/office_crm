import os
from pathlib import Path
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


COLOR_MAP = {
    "조건2_상병코드N회이상": "FFF9C4",
    "조건3_입원": "E0F2F1",
    "조건4_주상병특정코드": "BBDEFB",
    "조건5_총진료비이상": "FFE0B2",
    "조건6_총투약일수이상": "F8BBD0",
    "조건7_투약일수합산이상": "DCEDC8",
}

PALETTE = [
    "FFCDD2", "F8BBD0", "E1BEE7", "D1C4E9", "C5CAE9",
    "BBDEFB", "B3E5FC", "B2EBF2", "C8E6C9", "DCEDC8",
    "FFF9C4", "FFECB3", "FFE0B2", "FFCCBC"
]


def group_color(key):
    return PALETTE[hash(str(key)) % len(PALETTE)]


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = df.columns.astype(str)
    cols = cols.str.replace(r"x[0-9A-Fa-f]{4}", "", regex=True)
    cols = cols.str.replace(r"[\s\n\r\t]+", "", regex=True)
    cols = cols.str.replace(r"[^0-9a-zA-Z가-힣]", "", regex=True)
    cols = cols.str.strip()

    out = df.copy()
    out.columns = cols
    return out


def clean_text(value):
    if isinstance(value, str):
        return value.replace("_x000D_", "").strip()
    return value


def to_date_series(series):
    return pd.to_datetime(series, errors="coerce").dt.date


def fill_row_color(ws, row_idx: int, color: str):
    fill = PatternFill(
        start_color=color,
        end_color=color,
        fill_type="solid"
    )

    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def safe_save_workbook(wb, output_path: str) -> str:
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        base, ext = os.path.splitext(output_path)
        alt = f"{base}_{datetime.now().strftime('%H%M%S')}{ext}"
        wb.save(alt)
        return alt


def safe_read_excel(input_excel_path, sheet_index):
    try:
        return clean_columns(
            pd.read_excel(input_excel_path, sheet_name=sheet_index)
        )
    except Exception:
        return pd.DataFrame()


def build_summary_rows(df1_target, df3_target):
    rows = []

    for condition_name, df in {**df1_target, **df3_target}.items():
        rows.append({
            "조건": condition_name,
            "건수": len(df)
        })

    return pd.DataFrame(rows)


def analyze_excel_file(
    input_excel_path,
    output_dir,
    years=5,
    cond2_count=7,
    include_condition3=True,
    cond4_prefix_letters=None,
    cond4_second_letters=None,
    cond5_cost=150000,
    cond6_days=30,
    cond7_total_days=90,
    output_filename=None
):
    """
    변환된 심평원 엑셀 파일을 조건2~7 기준으로 분석합니다.

    기존 Analyze.py의 콘솔형 분석 흐름을 CRM에서 호출 가능한 함수로 바꾼 버전입니다.
    """

    input_excel_path = Path(input_excel_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not input_excel_path.exists():
        return {
            "success": False,
            "message": "입력 엑셀 파일을 찾을 수 없습니다.",
            "output_path": "",
            "summary": []
        }

    if cond4_prefix_letters is None:
        cond4_prefix_letters = ["F", "C", "D", "E", "G", "I", "J"]

    if cond4_second_letters is None:
        cond4_second_letters = ["F", "M", "C", "D", "E", "G", "I", "J"]

    cond4_prefix_letters = tuple(
        str(x).strip().upper()
        for x in cond4_prefix_letters
        if str(x).strip()
    )

    cond4_second_letters = tuple(
        str(x).strip().upper()
        for x in cond4_second_letters
        if str(x).strip()
    )

    cutoff = datetime.now().date() - timedelta(days=int(years) * 365)

    try:
        wb = load_workbook(input_excel_path)
    except Exception as e:
        return {
            "success": False,
            "message": f"엑셀 파일을 열 수 없습니다: {e}",
            "output_path": "",
            "summary": []
        }

    sheet_names = wb.sheetnames

    if len(sheet_names) < 3:
        return {
            "success": False,
            "message": f"시트가 3개 미만입니다. 현재 시트: {sheet_names}",
            "output_path": "",
            "summary": []
        }

    ws_base = wb[sheet_names[0]]
    ws_detail = wb[sheet_names[1]]

    df1 = safe_read_excel(input_excel_path, 0)
    df2 = safe_read_excel(input_excel_path, 1)
    df3 = safe_read_excel(input_excel_path, 2)

    if df1.empty and df2.empty and df3.empty:
        return {
            "success": False,
            "message": "엑셀에서 읽을 수 있는 데이터가 없습니다.",
            "output_path": "",
            "summary": []
        }

    for col in ["병의원약국", "진단과", "주상병명", "약품명", "처방조제"]:
        if col in df1.columns:
            df1[col] = df1[col].apply(clean_text)
        if col in df2.columns:
            df2[col] = df2[col].apply(clean_text)
        if col in df3.columns:
            df3[col] = df3[col].apply(clean_text)

    if "진료시작일" in df1.columns:
        df1["진료시작일"] = to_date_series(df1["진료시작일"])

    if "진료시작일" in df2.columns:
        df2["진료시작일"] = to_date_series(df2["진료시작일"])

    if "진료시작일" in df3.columns:
        df3["진료시작일"] = to_date_series(df3["진료시작일"])

    disease_code_col = "주상병코드" if "주상병코드" in df1.columns else None
    total_cost_col = "총진료비건강보험적용분" if "총진료비건강보험적용분" in df1.columns else None
    inout_col = "입원외래" if "입원외래" in df1.columns else None
    hospital_col = "병의원약국" if "병의원약국" in df1.columns else None

    total_days_col = "총투약일수" if "총투약일수" in df3.columns else None
    drug_name_col = "약품명" if "약품명" in df3.columns else None
    rx_type_col = "처방조제" if "처방조제" in df3.columns else None

    df1_5y = df1.copy()

    if "진료시작일" in df1_5y.columns:
        df1_5y = df1_5y[df1_5y["진료시작일"] >= cutoff]

    if "진단과" in df1_5y.columns:
        df1_5y = df1_5y[df1_5y["진단과"] != "일반의"]

    df1_target = {}
    df3_target = {}

    if disease_code_col:
        vc = df1_5y[disease_code_col].value_counts()
        codes = vc[vc >= int(cond2_count)].index

        cond_name = "조건2_상병코드N회이상"
        df1_target[cond_name] = df1_5y[
            df1_5y[disease_code_col].isin(codes)
        ].copy()

        sort_cols = [
            c for c in [disease_code_col, hospital_col, "진단과"]
            if c and c in df1_target[cond_name].columns
        ]

        if sort_cols:
            df1_target[cond_name] = df1_target[cond_name].sort_values(sort_cols)

    if include_condition3 and inout_col:
        df1_target["조건3_입원"] = df1_5y[
            df1_5y[inout_col] == "입원"
        ].copy()

    if disease_code_col:
        s = df1_5y[disease_code_col].astype(str).str.strip().str.upper()

        mask_prefix = s.str.startswith(cond4_prefix_letters, na=False)

        mask_second = (
            s.str.len().ge(2)
            & s.str[1].isin(cond4_second_letters)
        )

        cond_name = "조건4_주상병특정코드"
        df1_target[cond_name] = df1_5y[
            mask_prefix | mask_second
        ].copy()

        sort_cols = [
            c for c in [disease_code_col, hospital_col, "진단과"]
            if c and c in df1_target[cond_name].columns
        ]

        if sort_cols:
            df1_target[cond_name] = df1_target[cond_name].sort_values(sort_cols)

    if total_cost_col:
        tmp = df1_5y.copy()

        tmp[total_cost_col] = (
            tmp[total_cost_col]
            .astype(str)
            .str.replace(",", "")
            .str.strip()
        )

        tmp[total_cost_col] = (
            pd.to_numeric(tmp[total_cost_col], errors="coerce")
            .fillna(0)
            .astype(int)
        )

        df1_target["조건5_총진료비이상"] = tmp[
            tmp[total_cost_col] >= int(cond5_cost)
        ].copy()

    df3_work = df3.copy()

    if rx_type_col:
        df3_work = df3_work[
            df3_work[rx_type_col] != "외래"
        ].copy()

    if total_days_col:
        df3_work[total_days_col] = (
            pd.to_numeric(df3_work[total_days_col], errors="coerce")
            .fillna(0)
            .astype(int)
        )

    if total_days_col:
        df3_target["조건6_총투약일수이상"] = df3_work[
            df3_work[total_days_col] >= int(cond6_days)
        ].copy()

    if drug_name_col and total_days_col:
        g = df3_work.groupby(
            drug_name_col,
            dropna=False
        )[total_days_col].sum()

        hit_names = g[g >= int(cond7_total_days)].index

        df7 = df3_work[
            df3_work[drug_name_col].isin(hit_names)
        ].copy()

        sort_cols = [drug_name_col, "진료시작일"]
        sort_cols = [c for c in sort_cols if c in df7.columns]

        if sort_cols:
            df7 = df7.sort_values(sort_cols)

        df3_target["조건7_투약일수합산이상"] = df7

    if "분석" in wb.sheetnames:
        wb.remove(wb["분석"])

    ws = wb.create_sheet("분석")

    start_row = 1

    for cond_name, df_cond in {**df1_target, **df3_target}.items():
        ws.cell(row=start_row, column=1, value=cond_name)
        start_row += 1

        if df_cond.empty:
            ws.cell(row=start_row, column=1, value="해당 없음")
            start_row += 2
            continue

        df_ins = df_cond.copy()

        if "순번" not in df_ins.columns:
            df_ins.insert(0, "순번", range(1, len(df_ins) + 1))

        display_cols = ["순번"]

        if "진료시작일" in df_ins.columns:
            display_cols.append("진료시작일")

        other_cols = [
            c for c in df_ins.columns
            if c not in display_cols
        ]

        display_cols.extend(other_cols)
        df_ins = df_ins[display_cols]

        for r_idx, row in enumerate(
            dataframe_to_rows(df_ins, index=False, header=True),
            start=start_row
        ):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == start_row:
                continue

            if (
                cond_name in ["조건2_상병코드N회이상", "조건4_주상병특정코드"]
                and "주상병코드" in df_ins.columns
            ):
                code_val = df_ins.iloc[r_idx - start_row - 1]["주상병코드"]
                color = group_color(code_val)

            elif (
                cond_name == "조건7_투약일수합산이상"
                and drug_name_col in df_ins.columns
            ):
                name_val = df_ins.iloc[r_idx - start_row - 1][drug_name_col]
                color = group_color(name_val)

            else:
                color = COLOR_MAP.get(cond_name)

            if color:
                fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type="solid"
                )

                for fill_idx in range(1, len(row) + 1):
                    ws.cell(row=r_idx, column=fill_idx).fill = fill

        if cond_name in df1_target and hospital_col and "진료시작일" in df_ins.columns:
            base_color = COLOR_MAP.get(cond_name, "FFFFFF")

            for _, rr in df_ins.iterrows():
                d = rr.get("진료시작일")
                h = rr.get(hospital_col)

                if pd.isna(d) or pd.isna(h):
                    continue

                if "진료시작일" not in df1.columns or hospital_col not in df1.columns:
                    continue

                match_rows = df1[
                    (df1["진료시작일"] == d)
                    & (df1[hospital_col] == h)
                ].index

                for r in match_rows:
                    fill_row_color(ws_base, r + 2, base_color)

        if (
            cond_name == "조건5_총진료비이상"
            and hospital_col
            and "진료시작일" in df_ins.columns
        ):
            detail_color = COLOR_MAP.get(cond_name)

            for _, rr in df_ins.iterrows():
                d = rr.get("진료시작일")
                h = rr.get(hospital_col)

                if pd.isna(d) or pd.isna(h):
                    continue

                if "진료시작일" not in df2.columns or hospital_col not in df2.columns:
                    continue

                match_rows = df2[
                    (df2["진료시작일"] == d)
                    & (df2[hospital_col] == h)
                ].index

                for r in match_rows:
                    fill_row_color(ws_detail, r + 2, detail_color)

        start_row = ws.max_row + 2

    if "분석2" in wb.sheetnames:
        wb.remove(wb["분석2"])

    ws2 = wb.create_sheet("분석2")

    max_r = ws_detail.max_row
    max_c = ws_detail.max_column

    colored_rows = [
        [ws_detail.cell(row=1, column=c).value for c in range(1, max_c + 1)]
    ]

    for r in range(2, max_r + 1):
        has_fill = False

        for c in range(1, max_c + 1):
            fill = ws_detail.cell(row=r, column=c).fill

            if fill and fill.fill_type == "solid":
                has_fill = True
                break

        if has_fill:
            row_vals = [
                ws_detail.cell(row=r, column=c).value
                for c in range(1, max_c + 1)
            ]
            colored_rows.append(row_vals)

    for r_idx, row_vals in enumerate(colored_rows, start=1):
        for c_idx, value in enumerate(row_vals, start=1):
            ws2.cell(row=r_idx, column=c_idx, value=value)

    summary_df = build_summary_rows(df1_target, df3_target)

    summary = summary_df.to_dict("records")

    if output_filename:
        final_name = output_filename
        if not final_name.lower().endswith(".xlsx"):
            final_name += ".xlsx"
    else:
        final_name = f"{input_excel_path.stem}_분석파일.xlsx"

    output_path = output_dir / final_name

    try:
        saved_path = safe_save_workbook(wb, str(output_path))

    except Exception as e:
        return {
            "success": False,
            "message": f"분석 파일 저장 실패: {e}",
            "output_path": "",
            "summary": summary
        }

    return {
        "success": True,
        "message": "엑셀 분석이 완료되었습니다.",
        "output_path": saved_path,
        "summary": summary
    }