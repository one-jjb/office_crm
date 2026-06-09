# fill_master_with_mapping.py
# -*- coding: utf-8 -*-

import argparse
import datetime as dt
import difflib
import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter


# -----------------------------
# Data structures
# -----------------------------
@dataclass
class ExtractItem:
    run_id: str
    timestamp: str

    src_sheet: str
    src_row_in_sheet: int
    src_cell: str
    src_line_no_in_cell: int

    insurer: str
    product: str
    contract_date: str
    pay_term: str
    premium: str

    src_cat: str
    src_name: str
    src_amount_text: str
    src_status: str

    norm_cat: str
    match_text: str


@dataclass
class Rule:
    rule_id: int
    active: bool
    insurer: Optional[str]
    master_label: str
    target_field: str
    match_type: str
    pattern: str
    priority: int
    note: Optional[str]


# -----------------------------
# Constants
# -----------------------------
REPORT_HEADERS = [
    "insurer",
    "product",
    "src_cat",
    "src_name",
    "status",
    "final_master_label",
    "conflict",
]

CONFLICT_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
NO_FILL = PatternFill(fill_type=None)


# -----------------------------
# Helpers
# -----------------------------
def now_run_id() -> Tuple[str, str]:
    ts = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rid = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return rid, ts


def safe_str(x: Any) -> str:
    return "" if x is None else str(x)


def parse_amount_to_manh(amount_text: str) -> Optional[float]:
    s = safe_str(amount_text).strip()
    if not s:
        return None

    s = re.sub(r"\s+", "", s)
    sign = 1
    if s.startswith("-"):
        sign = -1
        s = s[1:]

    if s.endswith("만원"):
        num = s[:-2].replace(",", "")
        try:
            return sign * float(num)
        except Exception:
            return None

    if s.endswith("원"):
        num = s[:-1].replace(",", "")
        try:
            won = float(num)
            return sign * (won / 10000.0)
        except Exception:
            return None

    s2 = s.replace(",", "")
    try:
        return sign * float(s2)
    except Exception:
        return None


def display_amount_plain(amount_manh: float) -> str:
    val = float(amount_manh)
    return str(int(val)) if val.is_integer() else str(val)


def parse_contract_start_date(contract_date_text: str) -> Optional[dt.date]:
    s = safe_str(contract_date_text).strip()
    if not s:
        return None

    left = s.split("~")[0].strip()
    m = re.search(r"(\d{4})\.(\d{2})\.(\d{2})", left)
    if not m:
        return None

    try:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return dt.date(y, mo, d)
    except Exception:
        return None


def parse_multiline_cell_to_lines(cell_text: str) -> List[str]:
    if not cell_text:
        return []
    text = cell_text.replace("\r\n", "\n").replace("\r", "\n")
    return [ln.strip() for ln in text.split("\n") if ln.strip()]


def split_line_by_tabs(line: str) -> List[str]:
    parts = [p.strip() for p in line.split("\t")]
    while parts and parts[-1] == "":
        parts.pop()
    return parts


def extract_species_number(name: str) -> Optional[int]:
    m = re.search(r"\((\d)\s*종\)", name)
    if not m:
        return None
    try:
        n = int(m.group(1))
        return n if 1 <= n <= 5 else None
    except Exception:
        return None


def normalize_cat_for_1to5(cat: str, name: str) -> str:
    base = cat.strip()
    n = extract_species_number(name)
    if n is None:
        return base
    if base == "상해종수술":
        return f"상해 1-5종수술({n}종)"
    if base == "질병종수술":
        return f"질병 1-5종수술({n}종)"
    return base


def build_match_text(norm_cat: str, name: str) -> str:
    return f"{norm_cat} {name}".strip()


def get_or_create_sheet(wb: openpyxl.Workbook, title: str):
    if title in wb.sheetnames:
        return wb[title]
    return wb.create_sheet(title)


def clear_sheet(ws):
    ws.delete_rows(1, ws.max_row)


def build_comment_size(comment_text: str) -> Tuple[int, int]:
    lines = comment_text.split("\n") if comment_text else [""]
    max_len = max((len(line) for line in lines), default=1)
    line_count = max(len(lines), 1)

    width = max(150, min(1500, max_len * 14))
    height = max(70, min(1500, line_count * 26))
    return width, height


# -----------------------------
# Rules
# -----------------------------
def read_rules(ws) -> List[Rule]:
    headers = [safe_str(ws.cell(1, c).value).strip() for c in range(1, 60)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    required = ["rule_id", "active", "insurer", "master_label", "target_field", "match_type", "pattern", "priority"]
    for r in required:
        if r not in idx:
            raise ValueError(f"Mapping_Rules 헤더에 '{r}' 컬럼이 없습니다.")

    rules: List[Rule] = []
    for row in range(2, ws.max_row + 1):
        rid = ws.cell(row, idx["rule_id"]).value
        if rid is None:
            continue

        try:
            rule_id = int(rid)
        except Exception:
            continue

        active_raw = safe_str(ws.cell(row, idx["active"]).value).strip().lower()
        active = active_raw in ("y", "yes", "true", "1", "사용", "활성")

        insurer = safe_str(ws.cell(row, idx["insurer"]).value).strip()
        insurer = insurer if insurer else None

        master_label = safe_str(ws.cell(row, idx["master_label"]).value).strip()
        target_field = safe_str(ws.cell(row, idx["target_field"]).value).strip()
        match_type = safe_str(ws.cell(row, idx["match_type"]).value).strip()
        pattern = safe_str(ws.cell(row, idx["pattern"]).value)

        pr = ws.cell(row, idx["priority"]).value
        try:
            priority = int(pr) if pr is not None else 0
        except Exception:
            priority = 0

        note = safe_str(ws.cell(row, idx.get("note", 0)).value).strip() if "note" in idx else None

        rules.append(
            Rule(
                rule_id=rule_id,
                active=active,
                insurer=insurer,
                master_label=master_label,
                target_field=target_field,
                match_type=match_type,
                pattern=pattern,
                priority=priority,
                note=note,
            )
        )

    rules.sort(key=lambda x: (-x.priority, x.rule_id))
    return rules


def rule_matches(rule: Rule, insurer: str, norm_cat: str, name: str) -> bool:
    if not rule.active:
        return False
    if rule.insurer is not None and rule.insurer.strip() != insurer.strip():
        return False

    tf = rule.target_field.strip().lower()
    if tf == "cat+name":
        target_text = build_match_text(norm_cat, name)
    elif tf == "cat":
        target_text = norm_cat
    elif tf == "name":
        target_text = name
    else:
        return False

    mt = rule.match_type.strip().lower()
    pat = rule.pattern

    if mt == "keyword":
        keywords = [k.strip() for k in safe_str(pat).split("|") if k.strip()]
        if not keywords:
            return False
        return any(k in target_text for k in keywords)

    if mt == "regex":
        try:
            return re.search(pat, target_text) is not None
        except re.error:
            return False

    return False


def find_rule_for_item(rules: List[Rule], insurer: str, norm_cat: str, name: str) -> Optional[Rule]:
    for rule in rules:
        if rule_matches(rule, insurer, norm_cat, name):
            return rule
    return None


# -----------------------------
# Master helpers
# -----------------------------
def collect_master_label_rows(ws_master) -> Dict[str, int]:
    """
    Master_Data에서 마스터 담보 라벨이 있는 행을 찾는다.

    기존 코드는 Q열(17열)만 보았기 때문에 템플릿 구조가 조금만 달라도
    입력이 되지 않았다. 우선 Q열을 우선 사용하고, 없거나 부족하면
    A~Q 범위의 텍스트 셀을 함께 탐색한다.
    """
    out: Dict[str, int] = {}

    # 1순위: 기존 템플릿 기준 Q열
    for r in range(1, ws_master.max_row + 1):
        v = ws_master.cell(r, 17).value
        if isinstance(v, str) and v.strip():
            label = v.strip()
            if label not in ("리모델링 전", "리모델링 후", "비고"):
                out[label] = r

    # 2순위: 템플릿이 다를 경우 A~Q 범위에서 보장 라벨 후보 탐색
    for r in range(1, ws_master.max_row + 1):
        for c in range(1, min(ws_master.max_column, 17) + 1):
            v = ws_master.cell(r, c).value
            if not isinstance(v, str):
                continue
            label = v.strip()
            if not label:
                continue
            if label in ("보험사", "상품명", "계약일자", "납입기간", "보험료", "리모델링 전", "리모델링 후", "비고"):
                continue
            if len(label) > 60:
                continue
            # 너무 일반적인 표 제목/헤더는 제외
            if label in ("구분", "분류", "담보명", "가입금액", "합계", "소계"):
                continue
            out.setdefault(label, r)

    return out


def rank_candidates_with_score(master_labels: List[str], query: str, limit: int = 6) -> List[Tuple[str, float]]:
    query = query.strip()
    if not query:
        return [(lab, 0.0) for lab in master_labels[:limit]]

    scored: List[Tuple[str, float]] = []
    for lab in master_labels:
        ratio = difflib.SequenceMatcher(None, query, lab).ratio()
        scored.append((lab, ratio))
    scored.sort(key=lambda x: (-x[1], x[0]))
    return scored[:limit]


def read_kv_from_extract_sheet(ws) -> Dict[str, str]:
    keys = ("보험사", "상품명", "계약일자", "납입기간", "보험료")
    kv: Dict[str, str] = {}
    for r in range(1, min(ws.max_row, 80) + 1):
        key = safe_str(ws.cell(r, 2).value).strip()
        val = safe_str(ws.cell(r, 3).value).strip()
        if key in keys:
            kv[key] = val
    return kv


def fill_contract_header(ws_master, col: int, insurer: str, product: str, contract_date: str, pay_term: str, premium: str):
    ws_master.cell(5, col).value = insurer
    ws_master.cell(6, col).value = product
    ws_master.cell(8, col).value = contract_date
    ws_master.cell(9, col).value = pay_term
    if safe_str(ws_master.cell(10, col).value).strip() == "":
        ws_master.cell(10, col).value = premium


def clear_contract_header(ws_master, col: int):
    for r in (5, 6, 8, 9, 10):
        ws_master.cell(r, col).value = ""


def find_anchor_col(ws_master, text: str, header_row: int = 5) -> int:
    for c in range(1, ws_master.max_column + 1):
        v = ws_master.cell(header_row, c).value
        if isinstance(v, str) and text in v:
            return c
    raise ValueError(f"Master_Data {header_row}행에서 '{text}' 텍스트를 찾지 못했습니다.")


def parse_sum_range(formula: str) -> Optional[Tuple[int, int]]:
    if not isinstance(formula, str):
        return None
    m = re.search(r"SUM\(\s*([A-Z]+)\d+\s*:\s*([A-Z]+)\d+\s*\)", formula.replace("$", ""))
    if not m:
        return None
    try:
        s_col = column_index_from_string(m.group(1))
        e_col = column_index_from_string(m.group(2))
        return s_col, e_col
    except Exception:
        return None


def ensure_after_sum_formula(ws_master, after_anchor_col: int, new_start_col: int, current_end_col: int):
    cell = ws_master.cell(11, after_anchor_col)
    f = cell.value
    rng = parse_sum_range(f) if isinstance(f, str) else None
    if rng is None:
        return

    old_start, old_end = rng
    start_col = min(old_start, new_start_col)
    end_col = max(old_end, current_end_col)
    if start_col != old_start or end_col != old_end:
        cell.value = f"=SUM({get_column_letter(start_col)}11:{get_column_letter(end_col)}11)"


def find_remark_rows(ws_master, before_anchor_col: int, after_anchor_col: int) -> Tuple[int, int]:
    remark_cells: List[Tuple[int, int]] = []
    for r in range(1, ws_master.max_row + 1):
        for c in range(1, ws_master.max_column + 1):
            v = ws_master.cell(r, c).value
            if isinstance(v, str) and v.strip() == "비고":
                remark_cells.append((r, c))

    if not remark_cells:
        raise ValueError("Master_Data에서 '비고' 셀을 찾지 못했습니다.")

    def best_match(anchor_col: int, exclude: Optional[Tuple[int, int]] = None) -> Tuple[int, int]:
        candidates = [rc for rc in remark_cells if rc != exclude]
        if not candidates:
            raise ValueError("비고 기준 셀을 찾을 수 없습니다.")
        return min(candidates, key=lambda rc: (abs(rc[1] - anchor_col), rc[0], rc[1]))

    before_rc = best_match(before_anchor_col, exclude=None)
    after_rc = best_match(after_anchor_col, exclude=before_rc)

    return before_rc[0], after_rc[0]


# -----------------------------
# Extract
# -----------------------------
def extract_items_from_workbook(
    wb,
    run_id: str,
    timestamp: str,
    target_sheets: List[str],
) -> Tuple[List[ExtractItem], Dict[str, Dict[str, str]]]:
    items: List[ExtractItem] = []
    sheet_meta: Dict[str, Dict[str, str]] = {}

    for sname in target_sheets:
        if sname not in wb.sheetnames:
            continue

        ws = wb[sname]
        kv = read_kv_from_extract_sheet(ws)
        sheet_meta[sname] = kv

        insurer = kv.get("보험사", "").strip()
        product = kv.get("상품명", "").strip()
        contract_date = kv.get("계약일자", "").strip()
        pay_term = kv.get("납입기간", "").strip()
        premium = kv.get("보험료", "").strip()

        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row, col).value
                if not isinstance(val, str):
                    continue
                if "\t" not in val or ("\n" not in val and "\r" not in val):
                    continue

                cell_addr = f"{get_column_letter(col)}{row}"
                lines = parse_multiline_cell_to_lines(val)

                for line_no, line in enumerate(lines, start=1):
                    parts = split_line_by_tabs(line)
                    if len(parts) < 3:
                        continue

                    src_cat = parts[0]
                    src_name = parts[1]
                    src_amount = parts[2]
                    src_status = parts[3] if len(parts) >= 4 else ""

                    norm_cat = normalize_cat_for_1to5(src_cat, src_name)
                    match_text = build_match_text(norm_cat, src_name)

                    items.append(
                        ExtractItem(
                            run_id=run_id,
                            timestamp=timestamp,
                            src_sheet=sname,
                            src_row_in_sheet=row,
                            src_cell=cell_addr,
                            src_line_no_in_cell=line_no,
                            insurer=insurer,
                            product=product,
                            contract_date=contract_date,
                            pay_term=pay_term,
                            premium=premium,
                            src_cat=src_cat,
                            src_name=src_name,
                            src_amount_text=src_amount,
                            src_status=src_status,
                            norm_cat=norm_cat,
                            match_text=match_text,
                        )
                    )

    return items, sheet_meta


# -----------------------------
# Mapping_Report cache / merge
# -----------------------------
def read_mapping_report_cache(ws_report) -> Dict[Tuple[str, str, str], str]:
    headers = [safe_str(ws_report.cell(1, c).value).strip() for c in range(1, ws_report.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    required_basic = ["insurer", "src_cat", "src_name", "final_master_label"]
    if not all(k in idx for k in required_basic):
        return {}

    has_status = "status" in idx
    has_conflict = "conflict" in idx

    cache: Dict[Tuple[str, str, str], str] = {}
    for r in range(2, ws_report.max_row + 1):
        insurer = safe_str(ws_report.cell(r, idx["insurer"]).value).strip()
        src_cat = safe_str(ws_report.cell(r, idx["src_cat"]).value).strip()
        src_name = safe_str(ws_report.cell(r, idx["src_name"]).value).strip()
        final_label = safe_str(ws_report.cell(r, idx["final_master_label"]).value).strip()
        status = safe_str(ws_report.cell(r, idx["status"]).value).strip() if has_status else ""
        conflict = safe_str(ws_report.cell(r, idx["conflict"]).value).strip() if has_conflict else ""

        if not insurer or not src_cat or not src_name or not final_label:
            continue
        if conflict == "CONFLICT":
            continue
        if has_status and status and status not in (
            "MAPPED_BY_RULE",
            "MAPPED_BY_USER",
            "MAPPED_FROM_REPORT",
            "MAPPED_BY_100_MATCH",
        ):
            continue

        norm_cat = normalize_cat_for_1to5(src_cat, src_name)
        cache[(insurer, norm_cat, src_name)] = final_label

    return cache


def read_existing_mapping_report_rows(ws_report) -> List[Dict[str, str]]:
    headers = [safe_str(ws_report.cell(1, c).value).strip() for c in range(1, ws_report.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    required = ["insurer", "product", "src_cat", "src_name", "status", "final_master_label"]
    if not all(k in idx for k in required):
        return []

    rows: List[Dict[str, str]] = []
    for r in range(2, ws_report.max_row + 1):
        row = {
            "insurer": safe_str(ws_report.cell(r, idx["insurer"]).value).strip(),
            "product": safe_str(ws_report.cell(r, idx["product"]).value).strip(),
            "src_cat": safe_str(ws_report.cell(r, idx["src_cat"]).value).strip(),
            "src_name": safe_str(ws_report.cell(r, idx["src_name"]).value).strip(),
            "status": safe_str(ws_report.cell(r, idx["status"]).value).strip(),
            "final_master_label": safe_str(ws_report.cell(r, idx["final_master_label"]).value).strip(),
            "conflict": safe_str(ws_report.cell(r, idx["conflict"]).value).strip() if "conflict" in idx else "",
        }
        if any(row.values()):
            rows.append(row)

    return rows


def choose_preferred_status(a: str, b: str) -> str:
    priority = {
        "REMARK_BY_USER": 6,
        "MAPPED_BY_USER": 5,
        "MAPPED_FROM_REPORT": 4,
        "MAPPED_BY_RULE": 3,
        "MAPPED_BY_100_MATCH": 2,
        "UNMATCHED": 1,
        "": 0,
    }
    return a if priority.get(a, 0) >= priority.get(b, 0) else b


def merge_mapping_report_rows(
    existing_rows: List[Dict[str, str]],
    new_rows: List[Dict[str, str]],
) -> List[Dict[str, str]]:
    all_rows = existing_rows + new_rows

    grouped: Dict[Tuple[str, str, str, str], List[Dict[str, str]]] = {}
    for row in all_rows:
        key = (
            safe_str(row.get("insurer", "")).strip(),
            safe_str(row.get("product", "")).strip(),
            safe_str(row.get("src_cat", "")).strip(),
            safe_str(row.get("src_name", "")).strip(),
        )
        grouped.setdefault(key, []).append(row)

    merged_rows: List[Dict[str, str]] = []

    for key, rows in grouped.items():
        by_label: Dict[str, Dict[str, str]] = {}
        for row in rows:
            label = safe_str(row.get("final_master_label", "")).strip()
            status = safe_str(row.get("status", "")).strip()

            if label not in by_label:
                by_label[label] = {
                    "insurer": key[0],
                    "product": key[1],
                    "src_cat": key[2],
                    "src_name": key[3],
                    "status": status,
                    "final_master_label": label,
                    "conflict": "",
                }
            else:
                prev_status = safe_str(by_label[label].get("status", "")).strip()
                by_label[label]["status"] = choose_preferred_status(prev_status, status)

        unique_rows = list(by_label.values())
        non_empty = [r for r in unique_rows if safe_str(r.get("final_master_label", "")).strip()]

        if len(non_empty) == 0:
            best = unique_rows[0]
            best["conflict"] = ""
            merged_rows.append(best)
        elif len(non_empty) == 1:
            best = non_empty[0]
            best["conflict"] = ""
            merged_rows.append(best)
        else:
            for row in non_empty:
                row["conflict"] = "CONFLICT"
            merged_rows.extend(non_empty)

    merged_rows.sort(
        key=lambda x: (
            safe_str(x.get("insurer", "")),
            safe_str(x.get("product", "")),
            safe_str(x.get("src_cat", "")),
            safe_str(x.get("src_name", "")),
            safe_str(x.get("final_master_label", "")),
        )
    )
    return merged_rows


def write_mapping_report(ws_report, rows: List[Dict[str, Any]]):
    clear_sheet(ws_report)

    for c, h in enumerate(REPORT_HEADERS, start=1):
        ws_report.cell(1, c).value = h

    for r, rec in enumerate(rows, start=2):
        for c, h in enumerate(REPORT_HEADERS, start=1):
            ws_report.cell(r, c).value = rec.get(h, "")

    header_idx = {h: i + 1 for i, h in enumerate(REPORT_HEADERS)}
    col_final = header_idx["final_master_label"]
    col_conflict = header_idx["conflict"]

    for r in range(2, ws_report.max_row + 1):
        conflict = safe_str(ws_report.cell(r, col_conflict).value).strip()
        final_cell = ws_report.cell(r, col_final)
        conflict_cell = ws_report.cell(r, col_conflict)

        if conflict == "CONFLICT":
            final_cell.fill = CONFLICT_FILL
            conflict_cell.fill = CONFLICT_FILL
        else:
            final_cell.fill = NO_FILL
            conflict_cell.fill = NO_FILL


# -----------------------------
# Contract column assignment
# -----------------------------
def assign_contract_columns(ws_master, sheet_meta: Dict[str, Dict[str, str]]) -> Tuple[Dict[str, int], Dict[str, str]]:
    before_anchor = find_anchor_col(ws_master, "리모델링 전", header_row=5)
    after_anchor = find_anchor_col(ws_master, "리모델링 후", header_row=5)

    before_sum = parse_sum_range(ws_master.cell(11, before_anchor).value)
    after_sum = parse_sum_range(ws_master.cell(11, after_anchor).value)

    if before_sum is None:
        before_start = column_index_from_string("B")
        before_end = before_anchor - 1
    else:
        before_start, before_end = before_sum
        before_end = min(before_end, before_anchor - 1)

    if before_end < before_start:
        raise ValueError("리모델링 전 계약열 범위를 계산할 수 없습니다.")

    after_start_required = after_anchor + 1
    if after_sum is None:
        after_start = after_start_required
        after_end = ws_master.max_column
    else:
        _old_start, old_end = after_sum
        after_start = after_start_required
        after_end = old_end
        ensure_after_sum_formula(ws_master, after_anchor, new_start_col=after_start, current_end_col=after_end)

    if after_end < after_start:
        raise ValueError("리모델링 후 계약열 범위를 계산할 수 없습니다.")

    before_sheets: List[Tuple[str, dt.date]] = []
    after_sheets: List[str] = []

    for sname, kv in sheet_meta.items():
        cd = kv.get("계약일자", "").strip()
        sd = parse_contract_start_date(cd)
        if sd is not None:
            before_sheets.append((sname, sd))
        else:
            after_sheets.append(sname)

    before_sheets.sort(key=lambda x: x[1], reverse=True)

    before_cols_right_to_left = list(range(before_end, before_start - 1, -1))
    after_cols_right_to_left = list(range(after_end, after_start - 1, -1))

    if len(before_sheets) > len(before_cols_right_to_left):
        raise ValueError(f"리모델링 전 영역 열 부족: 필요 {len(before_sheets)} / 가능 {len(before_cols_right_to_left)}")
    if len(after_sheets) > len(after_cols_right_to_left):
        raise ValueError(f"리모델링 후 영역 열 부족: 필요 {len(after_sheets)} / 가능 {len(after_cols_right_to_left)}")

    for c in range(before_start, before_end + 1):
        clear_contract_header(ws_master, c)
    for c in range(after_start, after_end + 1):
        clear_contract_header(ws_master, c)

    sheet_to_col: Dict[str, int] = {}
    sheet_to_block: Dict[str, str] = {}

    for (sname, _sd), col in zip(before_sheets, before_cols_right_to_left):
        kv = sheet_meta[sname]
        fill_contract_header(
            ws_master, col,
            insurer=kv.get("보험사", "").strip(),
            product=kv.get("상품명", "").strip(),
            contract_date=kv.get("계약일자", "").strip(),
            pay_term=kv.get("납입기간", "").strip(),
            premium=kv.get("보험료", "").strip(),
        )
        sheet_to_col[sname] = col
        sheet_to_block[sname] = "BEFORE"

    for sname, col in zip(after_sheets, after_cols_right_to_left):
        kv = sheet_meta[sname]
        fill_contract_header(
            ws_master, col,
            insurer=kv.get("보험사", "").strip(),
            product=kv.get("상품명", "").strip(),
            contract_date=kv.get("계약일자", "").strip(),
            pay_term=kv.get("납입기간", "").strip(),
            premium=kv.get("보험료", "").strip(),
        )
        sheet_to_col[sname] = col
        sheet_to_block[sname] = "AFTER"

    return sheet_to_col, sheet_to_block


# -----------------------------
# Comment
# -----------------------------
def build_cell_comment_text(sources: List[str]) -> str:
    return "\n".join(sources)


# -----------------------------
# Main
# -----------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Input Excel (.xlsx)")
    ap.add_argument("--output", required=True, help="Output Excel (.xlsx)")
    ap.add_argument("--interactive", action="store_true", help="Ask user when unmapped")
    ap.add_argument("--write_comments", action="store_true", help="Write Excel cell comments")
    ap.add_argument("--auto_apply_100", action="store_true", help="Auto-apply when similarity is exactly 1.0")
    args = ap.parse_args()

    run_id, timestamp = now_run_id()
    wb = openpyxl.load_workbook(args.input)

    if "Mapping_Rules" not in wb.sheetnames:
        raise ValueError("입력 파일에 'Mapping_Rules' 시트가 없습니다.")
    master_sheet_name = None
    for candidate in ("Master_Data", "MASTER_DATA"):
        if candidate in wb.sheetnames:
            master_sheet_name = candidate
            break

    if master_sheet_name is None:
        raise ValueError("입력 파일에 'Master_Data' 또는 'MASTER_DATA' 시트가 없습니다.")

    ws_rules = wb["Mapping_Rules"]
    ws_master = wb[master_sheet_name]
    ws_report = get_or_create_sheet(wb, "Mapping_Report")

    rules = read_rules(ws_rules)
    master_label_to_row = collect_master_label_rows(ws_master)
    master_labels = list(master_label_to_row.keys())

    existing_report_rows: List[Dict[str, str]] = []
    report_cache: Dict[Tuple[str, str, str], str] = {}
    if "Mapping_Report" in wb.sheetnames and wb["Mapping_Report"].max_row >= 2:
        existing_report_rows = read_existing_mapping_report_rows(wb["Mapping_Report"])
        report_cache = read_mapping_report_cache(wb["Mapping_Report"])

    extract_sheet_names = [n for n in wb.sheetnames if n.startswith("Extract_Data")]
    items, sheet_meta = extract_items_from_workbook(wb, run_id, timestamp, extract_sheet_names)

    sheet_to_contract_col, sheet_to_block = assign_contract_columns(ws_master, sheet_meta)

    before_anchor_col = find_anchor_col(ws_master, "리모델링 전", header_row=5)
    after_anchor_col = find_anchor_col(ws_master, "리모델링 후", header_row=5)
    before_remark_row, after_remark_row = find_remark_rows(ws_master, before_anchor_col, after_anchor_col)

    new_report_rows: List[Dict[str, Any]] = []
    add_map: Dict[Tuple[int, int], float] = {}
    provenance_map: Dict[Tuple[int, int], List[str]] = {}
    remark_map: Dict[Tuple[int, int], List[str]] = {}

    for item in items:
        cached_label = report_cache.get((item.insurer.strip(), item.norm_cat.strip(), item.src_name.strip()))
        status = "UNMATCHED"
        final_label = ""

        if cached_label:
            final_label = cached_label
            status = "MAPPED_FROM_REPORT"
        else:
            matched_rule = find_rule_for_item(rules, item.insurer, item.norm_cat, item.src_name)
            if matched_rule is not None:
                final_label = matched_rule.master_label
                status = "MAPPED_BY_RULE"
            else:
                if args.auto_apply_100:
                    ranked = rank_candidates_with_score(master_labels, item.match_text, limit=1)
                    if ranked:
                        top_label, top_score = ranked[0]
                        if abs(top_score - 1.0) < 1e-12:
                            final_label = top_label
                            status = "MAPPED_BY_100_MATCH"

                if status == "UNMATCHED" and args.interactive:
                    candidates = rank_candidates_with_score(master_labels, item.match_text, limit=6)

                    print("\n[매핑 필요]")
                    print(f"- 시트명 : {item.src_sheet} / 보험사: {item.insurer}")
                    print(f"- 상품명 : {item.product}")
                    print(f"- 계약일자 : {item.contract_date}")
                    print(f"- 담보 : {item.src_cat} {item.src_name} {item.src_amount_text}  상태 : {item.src_status}")

                    print("\n매칭 값을 지정해주세요 :")
                    for i, (cand, score) in enumerate(candidates, start=1):
                        print(f"  {i}) {cand} ")
                    print("  0) 직접 입력")
                    print("  Enter) skip")
                    choice = input("선택: ").strip()

                    if choice == "":
                        status = "UNMATCHED"
                        final_label = ""
                    elif choice == "0":
                        typed = input("master_label 직접 입력: ").strip()
                        if typed == "":
                            status = "UNMATCHED"
                            final_label = ""
                        elif typed == "비고":
                            remark_text = input("비고 내용 입력: ").strip()
                            if remark_text:
                                remark_text = remark_text.replace("\\n", "\n")
                                contract_col = sheet_to_contract_col.get(item.src_sheet)
                                block = sheet_to_block.get(item.src_sheet, "")
                                if contract_col is not None:
                                    remark_row = before_remark_row if block == "BEFORE" else after_remark_row
                                    remark_map.setdefault((remark_row, contract_col), []).append(remark_text)
                                    final_label = "비고"
                                    status = "REMARK_BY_USER"
                                else:
                                    status = "UNMATCHED"
                                    final_label = ""
                            else:
                                status = "UNMATCHED"
                                final_label = ""
                        else:
                            final_label = typed
                            status = "MAPPED_BY_USER"
                    else:
                        try:
                            idx = int(choice)
                            if 1 <= idx <= len(candidates):
                                final_label = candidates[idx - 1][0]
                                status = "MAPPED_BY_USER"
                        except Exception:
                            if choice:
                                final_label = choice
                                status = "MAPPED_BY_USER"

        amount_manh = parse_amount_to_manh(item.src_amount_text)
        contract_col = sheet_to_contract_col.get(item.src_sheet)

        if (
            final_label
            and final_label != "비고"
            and final_label in master_label_to_row
            and contract_col is not None
            and amount_manh is not None
        ):
            r_master = master_label_to_row[final_label]
            key = (r_master, contract_col)
            add_map[key] = add_map.get(key, 0.0) + float(amount_manh)

            amount_disp = display_amount_plain(amount_manh)
            prov_line = f"{item.src_cat} {item.src_name} {amount_disp}"
            provenance_map.setdefault(key, []).append(prov_line)

        new_report_rows.append(
            {
                "insurer": item.insurer,
                "product": item.product,
                "src_cat": item.src_cat,
                "src_name": item.src_name,
                "status": status,
                "final_master_label": final_label,
                "conflict": "",
            }
        )

    comments_written = 0
    for (r, c), add_val in add_map.items():
        cell = ws_master.cell(r, c)

        current = cell.value
        try:
            cur_num = float(current) if current not in (None, "") else 0.0
        except Exception:
            cur_num = 0.0

        new_val = cur_num + add_val
        cell.value = int(round(new_val)) if abs(new_val - round(new_val)) < 1e-9 else new_val

        if args.write_comments:
            sources = provenance_map.get((r, c), [])
            comment_text = build_cell_comment_text(sources)

            comment = Comment(comment_text, "auto")
            width, height = build_comment_size(comment_text)
            comment.width = width
            comment.height = height

            cell.comment = comment
            comments_written += 1

    for (r, c), remark_lines in remark_map.items():
        cell = ws_master.cell(r, c)
        existing = safe_str(cell.value)

        if existing.strip():
            new_text = existing + "\n" + "\n".join(remark_lines)
        else:
            new_text = "\n".join(remark_lines)

        cell.value = new_text
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    merged_report_rows = merge_mapping_report_rows(existing_report_rows, new_report_rows)
    write_mapping_report(ws_report, merged_report_rows)

    wb.save(args.output)

    conflict_count = sum(1 for row in merged_report_rows if safe_str(row.get("conflict", "")).strip() == "CONFLICT")
    unmatched_count = sum(1 for row in merged_report_rows if safe_str(row.get("status", "")).strip() == "UNMATCHED")
    remark_count = sum(1 for row in merged_report_rows if safe_str(row.get("status", "")).strip() == "REMARK_BY_USER")

    print(f"\n완료: {args.output}")
    print(f"- 분석건수: {len(items)}")
    print(f"- 기존 Mapping_Report 행수: {len(existing_report_rows)}")
    print(f"- 이번 작업 신규 Mapping_Report 행수: {len(new_report_rows)}")
    print(f"- 누적 Mapping_Report 행수: {len(merged_report_rows)}")
    print(f"- UNMATCHED 행수: {unmatched_count}")
    print(f"- 비고 입력 행수: {remark_count}")
    print(f"- 충돌(CONFLICT) 행수: {conflict_count}")
    print(f"- Master_Data 입력 셀 수: {len(add_map)}")
    print(f"- 비고 입력 셀 수: {len(remark_map)}")
    print(f"- 메모 입력 셀 수: {comments_written} (enabled={args.write_comments})")


if __name__ == "__main__":
    main()