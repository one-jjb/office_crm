# C:\office_crm\views\coverage_excel_preview.py

import html
from pathlib import Path

import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


def color_to_hex(color):
    if not color:
        return None

    try:
        if color.type == "rgb" and color.rgb:
            value = color.rgb
            if len(value) == 8:
                return f"#{value[2:]}"
            if len(value) == 6:
                return f"#{value}"
    except Exception:
        return None

    return None


def cell_css(cell):
    css = []

    fill_color = color_to_hex(cell.fill.fgColor)
    if fill_color and fill_color.upper() != "#000000":
        css.append(f"background:{fill_color}")

    font_color = color_to_hex(cell.font.color)
    if font_color:
        css.append(f"color:{font_color}")

    if cell.font.bold:
        css.append("font-weight:700")

    if cell.font.sz:
        css.append(f"font-size:{int(cell.font.sz)}px")

    if cell.alignment.horizontal:
        css.append(f"text-align:{cell.alignment.horizontal}")

    if cell.alignment.vertical:
        css.append(f"vertical-align:{cell.alignment.vertical}")

    if cell.alignment.wrap_text:
        css.append("white-space:normal")
    else:
        css.append("white-space:nowrap")

    css.append("border:1px solid #222")
    css.append("padding:2px 4px")
    css.append("height:22px")
    css.append("box-sizing:border-box")

    return ";".join(css)


def get_merged_map(ws):
    merged_map = {}
    skip_cells = set()

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds

        merged_map[(min_row, min_col)] = {
            "rowspan": max_row - min_row + 1,
            "colspan": max_col - min_col + 1,
        }

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if not (row == min_row and col == min_col):
                    skip_cells.add((row, col))

    return merged_map, skip_cells


def worksheet_to_html(xlsx_path, max_rows=120, max_cols=45):
    xlsx_path = Path(xlsx_path)

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    merged_map, skip_cells = get_merged_map(ws)

    max_row = min(ws.max_row, max_rows)
    max_col = min(ws.max_column, max_cols)

    colgroup = []

    for col_idx in range(1, max_col + 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        width = ws.column_dimensions[letter].width or 10
        px = max(45, min(int(width * 7), 220))
        colgroup.append(f"<col style='width:{px}px'>")

    rows_html = []

    for row_idx in range(1, max_row + 1):
        row_cells = []

        for col_idx in range(1, max_col + 1):
            if (row_idx, col_idx) in skip_cells:
                continue

            cell = ws.cell(row=row_idx, column=col_idx)

            if isinstance(cell, MergedCell):
                continue

            value = "" if cell.value is None else str(cell.value)

            attrs = ""

            if (row_idx, col_idx) in merged_map:
                info = merged_map[(row_idx, col_idx)]

                if info["rowspan"] > 1:
                    attrs += f" rowspan='{info['rowspan']}'"

                if info["colspan"] > 1:
                    attrs += f" colspan='{info['colspan']}'"

            row_cells.append(
                f"<td{attrs} style='{cell_css(cell)}'>"
                f"{html.escape(value)}"
                f"</td>"
            )

        rows_html.append(f"<tr>{''.join(row_cells)}</tr>")

    return f"""
    <html>
    <head>
    <style>
        body {{
            margin:0;
            background:#f4f6f8;
            font-family:Malgun Gothic, Apple SD Gothic Neo, Arial, sans-serif;
        }}

        .excel-wrap {{
            width:100%;
            height:900px;
            overflow:auto;
            background:#f4f6f8;
            border:1px solid #d0d5dd;
            border-radius:10px;
        }}

        .excel-table {{
            border-collapse:collapse;
            table-layout:fixed;
            background:white;
            margin:12px;
            min-width:2200px;
        }}

        .excel-table td {{
            line-height:1.25;
        }}

        @media print {{
            body {{
                background:white;
            }}

            .excel-wrap {{
                height:auto;
                overflow:visible;
                border:none;
            }}

            .excel-table {{
                margin:0;
            }}
        }}
    </style>
    </head>
    <body>
        <div class="excel-wrap">
            <table class="excel-table">
                <colgroup>{''.join(colgroup)}</colgroup>
                <tbody>{''.join(rows_html)}</tbody>
            </table>
        </div>
    </body>
    </html>
    """


def render_excel_preview(xlsx_path):
    st.subheader("보장분석표 미리보기")

    html_view = worksheet_to_html(xlsx_path)
    components.html(html_view, height=930, scrolling=True)