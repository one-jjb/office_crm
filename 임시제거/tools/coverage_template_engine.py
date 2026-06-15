# C:\office_crm\tools\coverage_template_engine.py

from pathlib import Path
from datetime import datetime
import re
import shutil

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment


def clean_text(value):
    text = str(value or "")
    text = text.replace("_x000D_", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_text(value):
    text = clean_text(value)

    for ch in [
        " ", "\n", "\t", "\r",
        "-", "_",
        "(", ")", "[", "]",
        "/", "\\",
        ",", ".", "·", "ㆍ",
        ":",
    ]:
        text = text.replace(ch, "")

    return text.lower().strip()


def safe_filename(value):
    text = str(value or "").strip()

    for ch in r'\/:*?"<>|':
        text = text.replace(ch, "_")

    return text or "보장분석표.xlsx"


def split_mapping_labels(value):
    text = str(value or "").strip()

    if not text:
        return []

    parts = re.split(r"[|,\n;]+", text)

    labels = []
    seen = set()

    for part in parts:
        label = clean_text(part)

        if not label:
            continue

        if label in seen:
            continue

        seen.add(label)
        labels.append(label)

    return labels


def get_first_value(row, keys, default=""):
    for key in keys:
        if key in row and row.get(key) not in [None, ""]:
            return row.get(key)

    return default


def amount_to_number(value):
    text = clean_text(value).replace(",", "")

    if not text:
        return 0

    match = re.search(r"([\d.]+)\s*(억|천만|백만|만|원)?", text)

    if not match:
        return 0

    num = float(match.group(1))
    unit = match.group(2) or ""

    if unit == "억":
        return int(num * 100000000)
    if unit == "천만":
        return int(num * 10000000)
    if unit == "백만":
        return int(num * 1000000)
    if unit == "만":
        return int(num * 10000)

    return int(num)


def format_amount(value):
    text = clean_text(value)

    if not text:
        return ""

    if re.search(r"[가-힣]", text):
        return text

    num = amount_to_number(text)

    if not num:
        return text

    if num % 10000 == 0:
        return f"{num // 10000:,}만"

    return f"{num:,}"


def expand_multi_mapping_rows(mapping_df):
    if mapping_df is None or mapping_df.empty:
        return pd.DataFrame()

    expanded_rows = []

    for _, row in mapping_df.iterrows():
        row_dict = row.to_dict()

        target_value = get_first_value(
            row_dict,
            ["확정위치", "target_label", "매핑대상", "추천위치"],
        )

        labels = split_mapping_labels(target_value)

        if not labels:
            expanded_rows.append(row_dict)
            continue

        for label in labels:
            new_row = row_dict.copy()
            new_row["확정위치"] = label
            expanded_rows.append(new_row)

    return pd.DataFrame(expanded_rows).reset_index(drop=True)


def find_sheet_by_priority(workbook):
    preferred_names = [
        "보장분석",
        "보장분석표",
        "MASTER_DATA",
        "Master",
        "master",
        "Sheet1",
    ]

    for name in preferred_names:
        if name in workbook.sheetnames:
            return workbook[name]

    return workbook[workbook.sheetnames[0]]


def find_marker_cell(ws, markers):
    marker_norms = [normalize_text(marker) for marker in markers]

    for row in ws.iter_rows():
        for cell in row:
            value_norm = normalize_text(cell.value)

            if not value_norm:
                continue

            for marker_norm in marker_norms:
                if marker_norm and marker_norm in value_norm:
                    return cell

    return None


def find_remodeling_start_col(ws, remodel_type):
    if remodel_type == "리모델링 후":
        markers = ["리모델링 후", "리모델링후", "변경 후", "변경후"]
        direction = 1
    else:
        markers = ["리모델링 전", "리모델링전", "변경 전", "변경전"]
        direction = -1

    marker_cell = find_marker_cell(ws, markers)

    if marker_cell:
        if remodel_type == "리모델링 후":
            return marker_cell.column + 1
        return marker_cell.column - 1 if marker_cell.column > 1 else marker_cell.column

    return None


def find_target_label_cell(ws, target_label):
    target_norm = normalize_text(target_label)

    if not target_norm:
        return None

    best_cell = None
    best_score = 0

    for row in ws.iter_rows():
        for cell in row:
            value = clean_text(cell.value)
            value_norm = normalize_text(value)

            if not value_norm:
                continue

            score = 0

            if value_norm == target_norm:
                score = 100
            elif target_norm in value_norm:
                score = 90
            elif value_norm in target_norm:
                score = 80

            if score > best_score:
                best_score = score
                best_cell = cell

    if best_score >= 80:
        return best_cell

    return None


def find_output_cell(ws, target_label, remodel_type):
    label_cell = find_target_label_cell(ws, target_label)

    if not label_cell:
        return None

    start_col = find_remodeling_start_col(ws, remodel_type)

    if start_col:
        if remodel_type == "리모델링 후":
            output_col = max(start_col, label_cell.column + 1)
        else:
            output_col = start_col

            if output_col == label_cell.column:
                output_col = label_cell.column + 1
    else:
        output_col = label_cell.column + 1

    if output_col < 1:
        output_col = label_cell.column + 1

    return ws.cell(row=label_cell.row, column=output_col)


def build_comment_text(rows):
    lines = []

    for row in rows:
        company = get_first_value(row, ["보험사", "company"], "")
        product = get_first_value(row, ["상품명", "product"], "")
        company_cov = get_first_value(row, ["회사담보명", "company_coverage_name"], "")
        credit_cov = get_first_value(row, ["신정원담보명", "credit_coverage_name"], "")
        source = get_first_value(row, ["추출담보명", "source_name"], "")
        amount = get_first_value(row, ["보장금액", "amount"], "")

        line = " / ".join(
            [
                item
                for item in [
                    company,
                    product,
                    company_cov or source,
                    credit_cov,
                    amount,
                ]
                if clean_text(item)
            ]
        )

        if line:
            lines.append(line)

    return "\n".join(lines)


def resize_comment(comment, text):
    if not comment:
        return

    lines = str(text or "").splitlines()
    max_len = max([len(line) for line in lines], default=10)
    line_count = max(len(lines), 1)

    comment.width = max(150, min(1500, max_len * 14))
    comment.height = max(70, min(1500, line_count * 26))


def write_value_with_comment(cell, amount_value, comment_text):
    existing_value = clean_text(cell.value)
    new_value = clean_text(amount_value)

    if existing_value and existing_value != new_value:
        cell.value = f"{existing_value} / {new_value}"
    else:
        cell.value = new_value

    if comment_text:
        comment = Comment(comment_text, "office_crm")
        resize_comment(comment, comment_text)
        cell.comment = comment


def group_mapping_rows(mapping_df):
    grouped = {}

    for _, row in mapping_df.iterrows():
        row_dict = row.to_dict()

        status = get_first_value(row_dict, ["상태", "status"], "")
        target_label = get_first_value(row_dict, ["확정위치", "target_label"], "")

        if status == "매핑 제외":
            continue

        if not target_label or target_label == "매핑 제외":
            continue

        amount = get_first_value(row_dict, ["보장금액", "amount"], "")

        key = target_label

        if key not in grouped:
            grouped[key] = {
                "target_label": target_label,
                "amount_total": 0,
                "amount_values": [],
                "rows": [],
            }

        amount_number = amount_to_number(amount)

        if amount_number:
            grouped[key]["amount_total"] += amount_number

        if amount:
            grouped[key]["amount_values"].append(amount)

        grouped[key]["rows"].append(row_dict)

    return grouped


def decide_display_amount(group):
    amount_total = group.get("amount_total", 0)
    amount_values = group.get("amount_values", [])

    if amount_total:
        if amount_total % 10000 == 0:
            return f"{amount_total // 10000:,}만"
        return f"{amount_total:,}"

    unique_values = []

    for value in amount_values:
        value = clean_text(value)
        if value and value not in unique_values:
            unique_values.append(value)

    return " / ".join(unique_values)


def safe_save_workbook(wb, output_path):
    output_path = Path(output_path)

    try:
        wb.save(output_path)
        return output_path

    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt_path = output_path.with_name(
            f"{output_path.stem}_{timestamp}{output_path.suffix}"
        )
        wb.save(alt_path)
        return alt_path


def generate_coverage_report_from_mapping(
    template_path,
    mapping_df,
    output_dir,
    output_filename,
    remodel_type="리모델링 전",
):
    template_path = Path(template_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    output_filename = safe_filename(output_filename)

    if not output_filename.lower().endswith(".xlsx"):
        output_filename += ".xlsx"

    output_path = output_dir / output_filename

    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path)
    ws = find_sheet_by_priority(wb)

    mapping_df = expand_multi_mapping_rows(mapping_df)
    grouped = group_mapping_rows(mapping_df)

    written = []
    unmatched = []

    for target_label, group in grouped.items():
        output_cell = find_output_cell(
            ws=ws,
            target_label=target_label,
            remodel_type=remodel_type,
        )

        if not output_cell:
            unmatched.append(
                {
                    "확정위치": target_label,
                    "사유": "템플릿에서 대상 라벨을 찾지 못함",
                    "담보수": len(group.get("rows", [])),
                    "보장금액": decide_display_amount(group),
                }
            )
            continue

        amount_value = decide_display_amount(group)
        comment_text = build_comment_text(group.get("rows", []))

        write_value_with_comment(
            cell=output_cell,
            amount_value=amount_value,
            comment_text=comment_text,
        )

        written.append(
            {
                "확정위치": target_label,
                "입력셀": output_cell.coordinate,
                "보장금액": amount_value,
                "담보수": len(group.get("rows", [])),
            }
        )

    final_output_path = safe_save_workbook(wb, output_path)

    return {
        "output_path": str(final_output_path),
        "written": written,
        "unmatched": unmatched,
    }