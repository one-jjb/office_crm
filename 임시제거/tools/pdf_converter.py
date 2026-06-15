import os
import sys
import time
import re
import unicodedata
from pathlib import Path
from copy import copy

import pikepdf
import tabula
import pandas as pd
import pdfplumber
import jpype

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


# =========================================================
# JVM / Java 설정
# =========================================================
def get_base_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent

    return Path(__file__).resolve().parent.parent


def find_jvm_path():
    base_dir = get_base_dir()

    candidate_paths = [
        base_dir / "jre" / "bin" / "server" / "jvm.dll",
        base_dir / "_internal" / "jre" / "bin" / "server" / "jvm.dll",
        Path(r"C:\PDF_CONVERT\dist\PDF_Conversion\_internal\jre\bin\server\jvm.dll"),
        Path(r"C:\PDF_CONVERT\jre\bin\server\jvm.dll"),
    ]

    for path in candidate_paths:
        if path.exists():
            return path

    return None


def ensure_jvm_started():
    if jpype.isJVMStarted():
        return

    jvm_path = find_jvm_path()

    if not jvm_path:
        raise FileNotFoundError(
            "JVM 파일을 찾을 수 없습니다. "
            "jre\\bin\\server\\jvm.dll 위치를 확인하세요."
        )

    try:
        jpype.startJVM(
            str(jvm_path),
            "-ea",
            "--enable-native-access=ALL-UNNAMED",
            convertStrings=True
        )
    except Exception as e:
        raise RuntimeError(f"JVM 시작 실패: {e}")


# =========================================================
# 기본 유틸
# =========================================================
def ts_name():
    return time.strftime("%y_%m_%d_%H_%M_%S")


def safe_filename(name: str) -> str:
    name = str(name)
    name = re.sub(r'[\\/:*?"<>|]', "_", name)
    name = name.strip()

    if not name:
        return "file"

    return name


def sanitize_sheet_name(filename: str) -> str:
    base = os.path.splitext(os.path.basename(filename))[0]

    match = re.search(r"\((.*?)\)", base)

    if match:
        name = match.group(1)
    else:
        name = base

    forbidden = "[]:*?/\\"

    name = "".join("_" if ch in forbidden else ch for ch in name).strip()

    if not name:
        name = "Sheet"

    return name[:31]


def unlock_pdf(input_pdf: str, password: str, output_pdf: str) -> bool:
    try:
        with pikepdf.open(input_pdf, password=password) as pdf:
            pdf.save(output_pdf)

        return True

    except Exception as e:
        raise RuntimeError(
            f"PDF 잠금 해제 실패: {os.path.basename(input_pdf)} / {e}"
        )


def read_tables_with_tabula(pdf_path: str):
    ensure_jvm_started()

    try:
        dfs = tabula.read_pdf(
            pdf_path,
            pages="all",
            multiple_tables=True
        )

        if dfs:
            return dfs

    except Exception as e:
        print(f"[안내] tabula 실패: {e}")

    return []


def read_text_with_pdfplumber(pdf_path: str) -> list:
    texts = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                txt = page.extract_text() or ""

                if txt:
                    txt = "\n".join(
                        line.strip()
                        for line in txt.splitlines()
                    )

                texts.append(txt)

    except Exception as e:
        print(f"[안내] pdfplumber 실패: {e}")

    return texts


def normalize_extracted_text(value):
    if value is None:
        return None

    text = str(value)

    text = (
        text.replace("\r\n", "")
        .replace("\n", "")
        .replace("\r", "")
        .replace("\t", "")
    )

    text = re.sub(r"\s{2,}", " ", text)

    return text.strip()


def excel_display_width(value: str) -> int:
    if value is None:
        return 0

    text = str(value)
    width = 0

    for ch in text:
        ea = unicodedata.east_asian_width(ch)
        width += 2 if ea in ("W", "F") else 1

    return width


def autofit_sheet_columns(
    ws,
    max_width: int = 80,
    min_width: int = 8,
    padding: int = 2
):
    no_wrap = Alignment(wrap_text=False)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            if cell.alignment:
                new_align = copy(cell.alignment)
                new_align.wrap_text = False
                cell.alignment = new_align
            else:
                cell.alignment = no_wrap

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0

        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value

            if value is None:
                continue

            text = (
                str(value)
                .replace("\r\n", " ")
                .replace("\n", " ")
                .replace("\r", " ")
            )

            max_len = max(
                max_len,
                excel_display_width(text)
            )

        width = max(
            min_width,
            min(max_width, max_len + padding)
        )

        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_tables_to_sheet(ws, dfs: list) -> bool:
    if not dfs:
        return False

    base_columns = None
    wrote_any = False

    for df in dfs:
        if not isinstance(df, pd.DataFrame) or df.empty:
            continue

        df = df.copy()

        df.columns = [
            normalize_extracted_text(col)
            for col in df.columns
        ]

        try:
            df = df.map(normalize_extracted_text)
        except AttributeError:
            df = df.applymap(normalize_extracted_text)

        if base_columns is None:
            base_columns = [
                normalize_extracted_text(col)
                for col in list(df.columns)
            ]

            ws.append(base_columns)

        else:
            if list(df.columns) != base_columns:
                continue

        first_row = [
            normalize_extracted_text(value)
            for value in df.iloc[0].tolist()
        ]

        header_row = [
            normalize_extracted_text(col)
            for col in base_columns
        ]

        if first_row == header_row:
            df = df.iloc[1:]

        for row in df.itertuples(index=False):
            ws.append(list(row))

        wrote_any = True

    return wrote_any


def write_text_to_sheet(ws, texts: list) -> bool:
    if not texts:
        return False

    for i, txt in enumerate(texts, start=1):
        ws.append([f"페이지 {i}"])

        if txt:
            for line in txt.splitlines():
                ws.append([line])
        else:
            ws.append(["(텍스트 없음)"])

        ws.append([])

    return True


# =========================================================
# CRM 호출용 메인 함수
# =========================================================
def convert_pdfs_to_excel(
    pdf_paths,
    password,
    output_dir,
    output_filename=None
):
    """
    CRM/Streamlit에서 호출할 함수.

    Parameters
    ----------
    pdf_paths : list[str | Path]
        변환할 PDF 파일 경로 목록.
    password : str
        PDF 비밀번호.
    output_dir : str | Path
        결과 엑셀 저장 폴더.
    output_filename : str | None
        결과 파일명. None이면 자동 생성.

    Returns
    -------
    dict
        {
            "success": bool,
            "output_path": str,
            "converted_count": int,
            "failed_files": list[str],
            "message": str
        }
    """

    if not pdf_paths:
        return {
            "success": False,
            "output_path": "",
            "converted_count": 0,
            "failed_files": [],
            "message": "업로드된 PDF 파일이 없습니다."
        }

    if not password:
        return {
            "success": False,
            "output_path": "",
            "converted_count": 0,
            "failed_files": [],
            "message": "PDF 비밀번호가 입력되지 않았습니다."
        }

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    temp_unlock_dir = output_dir / "_tmp_unlocked"
    temp_unlock_dir.mkdir(parents=True, exist_ok=True)

    if output_filename:
        final_name = safe_filename(output_filename)
        if not final_name.lower().endswith(".xlsx"):
            final_name += ".xlsx"
    else:
        final_name = f"converted_{ts_name()}.xlsx"

    final_path = output_dir / final_name

    wb = Workbook()
    wb.remove(wb.active)

    created_any_sheet = False
    converted_count = 0
    failed_files = []

    for pdf_path in pdf_paths:
        pdf_path = Path(pdf_path)

        if not pdf_path.exists():
            failed_files.append(f"{pdf_path.name} / 파일 없음")
            continue

        unlocked_path = temp_unlock_dir / f"unlocked_{pdf_path.name}"

        try:
            unlock_pdf(
                str(pdf_path),
                password,
                str(unlocked_path)
            )

            sheet_name = sanitize_sheet_name(pdf_path.name)
            ws = wb.create_sheet(title=sheet_name)

            dfs = read_tables_with_tabula(str(unlocked_path))
            wrote_tables = write_tables_to_sheet(ws, dfs)

            if wrote_tables:
                created_any_sheet = True
            else:
                texts = read_text_with_pdfplumber(str(unlocked_path))

                if write_text_to_sheet(ws, texts):
                    created_any_sheet = True
                else:
                    ws.append(["정보", "표/텍스트 추출 실패"])

            try:
                autofit_sheet_columns(ws)
            except Exception as e:
                print(f"[안내] 열너비 자동조절 실패({sheet_name}): {e}")

            converted_count += 1

        except Exception as e:
            failed_files.append(f"{pdf_path.name} / {e}")

    if not created_any_sheet:
        ws = wb.create_sheet(title="정보")
        ws.append(["정보", "모든 PDF에서 데이터 추출 실패"])

    try:
        wb.save(final_path)

    except Exception as e:
        return {
            "success": False,
            "output_path": "",
            "converted_count": converted_count,
            "failed_files": failed_files,
            "message": f"엑셀 저장 실패: {e}"
        }

    return {
        "success": True,
        "output_path": str(final_path),
        "converted_count": converted_count,
        "failed_files": failed_files,
        "message": f"PDF {converted_count}개 변환 완료"
    }


# =========================================================
# 단독 테스트용
# =========================================================
if __name__ == "__main__":
    test_input_dir = Path("Input_PDFs")
    test_output_dir = Path("Output")

    pdf_files = sorted(test_input_dir.glob("*.pdf"))

    password = input("PDF 비밀번호를 입력하세요: ").strip()

    result = convert_pdfs_to_excel(
        pdf_paths=pdf_files,
        password=password,
        output_dir=test_output_dir
    )

    print(result)